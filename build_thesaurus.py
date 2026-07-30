import collections
import csv
import glob
import json
import math
import os
import re

import numpy as np
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CORPORA = os.environ.get(
    "CORPORA_DIR", os.path.join(ROOT, "dict_and_corpora", "corpora")
)
DATA = os.environ.get("APP_DATA_DIR", os.path.join(ROOT, "data"))
OUT = os.path.join(DATA, "huayu", "thesaurus.json")
CACHE = os.path.join(ROOT, "dict_and_corpora", "thesaurus_tokens.npz")

HAN = re.compile(r"^[㐀-鿿]+$")
MARK = re.compile(r"\[[似反]\]")
SENSE_NUMBER = re.compile(r"^\s*\d+\s*[.、]\s*")
SPLIT = re.compile(r"[、，,；;／/\s]+")

MAX_WORD = 6
MIN_TARGET_FREQ = 25
WINDOW = 2
ALPHA = 0.75
SHIFT = 2.0
MIN_PAIR_COUNT = 3
TOP_CONTEXTS = 300
NEIGHBORS = 10
MIN_COSINE = 0.22
MUTUAL_RANK = 100
COLLOCATE_MIN_PARTNER_FREQ = 50
SVD_DIMENSIONS = 300
SVD_POWER = 0.5
SVD_OVERSAMPLING = 10
SVD_ITERATIONS = 2
SVD_SEED = 20260724

COLLOCATE_WINDOW = 3
COLLOCATE_MIN_COUNT = 8
COLLOCATE_MIN_G2 = 10.83
COLLOCATES = 8

FUNCTION_POS = {"Ptc", "Det", "Prep", "Conj"}
FUNCTION_WORDS = set(
    "的了是在有和與或而但就也都又並把被將給對從跟為以之其此該每各另"
    "我你妳他她它牠祂們這那哪誰甚麼麼呢嗎吧啊喔耶欸"
) | {
    "什麼",
    "為什麼",
    "怎麼",
    "這樣",
    "那樣",
    "沒有",
    "可以",
    "因為",
    "所以",
    "但是",
    "如果",
    "而且",
    "已經",
    "還有",
    "這個",
    "那個",
    "一個",
    "一些",
    "我們",
    "你們",
    "他們",
    "她們",
    "自己",
    "大家",
    "時候",
    "現在",
}

def read_xlsx(name):
    path = os.path.join(CORPORA, "moedict", name)
    book = openpyxl.load_workbook(path, read_only=True)
    sheet = book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index = {title: position for position, title in enumerate(header)}
    for row in rows:
        yield row, index
    book.close()

def parse_list(cell):
    if not cell:
        return []
    text = MARK.sub("", str(cell))
    out = []
    for part in SPLIT.split(text):
        part = SENSE_NUMBER.sub("", part).strip("「」『』（）()。 ")
        part = SENSE_NUMBER.sub("", part)
        if part and HAN.match(part):
            out.append(part)
    return out

def dictionary_relations():
    synonyms = collections.defaultdict(set)
    antonyms = collections.defaultdict(set)
    for name in ("dict_concised.xlsx", "dict_revised.xlsx"):
        for row, index in read_xlsx(name):
            word = row[index["字詞名"]]
            if not word:
                continue
            word = str(word)
            synonyms[word].update(parse_list(row[index["相似詞"]]))
            antonyms[word].update(parse_list(row[index["相反詞"]]))

    for store in (synonyms, antonyms):
        for word, partners in list(store.items()):
            for partner in partners:
                store[partner].add(word)

    return (
        {
            word: sorted(partners - {word})
            for word, partners in synonyms.items()
            if partners - {word}
        },
        {
            word: sorted(partners - {word})
            for word, partners in antonyms.items()
            if partners - {word}
        },
    )

def dictionary_words():
    words = set()
    with open(os.path.join(CORPORA, "concised.json"), encoding="utf-8") as handle:
        words.update(entry["word"] for entry in json.load(handle))
    with open(
        os.path.join(DATA, "huayu", "school_levels.json"), encoding="utf-8"
    ) as handle:
        words.update(row["traditional"] for row in json.load(handle))
    with open(os.path.join(DATA, "huayu", "tocfl.csv"), encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            raw = row.get("Variants") or ""
            forms = []
            if raw.strip():
                try:
                    forms = [entry["Traditional"].strip() for entry in json.loads(raw)]
                except json.JSONDecodeError:
                    forms = []
            words.update(
                forms
                or [form.strip() for form in re.split(r"[/／]", row["Traditional"])]
            )
    return {word for word in words if word and HAN.match(word)}

def segment(words):
    prefixes = set()
    for word in words:
        for size in range(2, len(word)):
            prefixes.add(word[:size])

    tokens = []
    add = tokens.append
    for path in sorted(glob.glob(os.path.join(DATA, "corpora", "sentences", "*.json"))):
        with open(path, encoding="utf-8") as handle:
            for sentence in json.load(handle):
                position = 0
                length = len(sentence)
                while position < length:
                    char = sentence[position]
                    if not ("㐀" <= char <= "鿿"):
                        position += 1
                        add(None)
                        continue
                    step = 1
                    token = char
                    if (
                        sentence[position : position + 2] in words
                        or sentence[position : position + 2] in prefixes
                    ):
                        for size in range(min(MAX_WORD, length - position), 1, -1):
                            if sentence[position : position + size] in words:
                                token = sentence[position : position + size]
                                step = size
                                break
                    add(token)
                    position += step
                add(None)
    return tokens

def cached_stream(words):
    if os.path.exists(CACHE):
        stored = np.load(CACHE, allow_pickle=False)
        return stored["stream"], list(stored["tokens"])

    tokens = segment(words)
    alphabet = sorted({token for token in tokens if token})
    index = {token: number for number, token in enumerate(alphabet)}
    stream = np.fromiter(
        (index.get(token, -1) for token in tokens), dtype=np.int32, count=len(tokens)
    )
    np.savez_compressed(CACHE, stream=stream, tokens=np.array(alphabet))
    return stream, alphabet

def restrict(stream, alphabet, keep):
    mapping = np.full(len(alphabet) + 1, -1, dtype=np.int32)
    for number, token in enumerate(alphabet):
        position = keep.get(token)
        if position is not None:
            mapping[number] = position
    shifted = np.where(stream >= 0, stream, len(alphabet))
    return mapping[shifted]

def pair_counts(stream, size, window, symmetric=True):
    keys = []
    counts = []
    for distance in range(1, window + 1):
        left = stream[:-distance]
        right = stream[distance:]
        keep = (left >= 0) & (right >= 0)
        if not keep.any():
            continue
        first = left[keep].astype(np.int64)
        second = right[keep].astype(np.int64)
        both = (
            np.concatenate([first * size + second, second * size + first])
            if symmetric
            else first * size + second
        )
        unique, count = np.unique(both, return_counts=True)
        keys.append(unique)
        counts.append(count)

    merged = np.concatenate(keys)
    weights = np.concatenate(counts)
    order = np.argsort(merged, kind="stable")
    merged = merged[order]
    weights = weights[order]
    unique, start = np.unique(merged, return_index=True)
    total = np.add.reduceat(weights, start)
    return unique, total

def sppmi_rows(keys, counts, size, alpha, shift, min_count):
    keep = counts >= min_count
    keys = keys[keep]
    counts = counts[keep].astype(np.float64)
    targets = keys // size
    contexts = keys % size

    row_totals = np.bincount(targets, weights=counts, minlength=size)
    context_totals = np.bincount(contexts, weights=counts, minlength=size)
    smoothed = context_totals**alpha
    smoothed_total = smoothed.sum()
    grand_total = counts.sum()

    with np.errstate(divide="ignore", invalid="ignore"):
        value = np.log(
            counts * smoothed_total / (row_totals[targets] * smoothed[contexts])
        )
    value -= math.log(shift) if shift > 1 else 0.0

    positive = value > 0
    return targets[positive], contexts[positive], value[positive]

def build_space(targets, contexts, values, size, top_contexts):
    matrix = np.zeros((size, size), dtype=np.float32)
    order = np.lexsort((-values, targets))
    targets = targets[order]
    contexts = contexts[order]
    values = values[order]

    boundaries = np.flatnonzero(np.diff(targets)) + 1
    for start, stop in zip(
        np.concatenate([[0], boundaries]), np.concatenate([boundaries, [len(targets)]])
    ):
        stop = min(stop, start + top_contexts)
        matrix[targets[start], contexts[start:stop]] = values[start:stop]

    norms = np.linalg.norm(matrix, axis=1, keepdims=True)
    np.divide(matrix, norms, out=matrix, where=norms > 0)
    return matrix

def randomized_svd(
    matrix, dimensions, oversampling=SVD_OVERSAMPLING, iterations=SVD_ITERATIONS
):
    generator = np.random.default_rng(SVD_SEED)
    probe = generator.standard_normal(
        size=(matrix.shape[1], dimensions + oversampling), dtype=np.float32
    )
    sample = matrix @ probe
    basis, _ = np.linalg.qr(sample)
    for _ in range(iterations):
        basis, _ = np.linalg.qr(matrix.T @ basis)
        basis, _ = np.linalg.qr(matrix @ basis)

    small = basis.T @ matrix
    left, singular, _ = np.linalg.svd(small, full_matrices=False)
    return (basis @ left)[:, :dimensions], singular[:dimensions]

def embed(matrix, dimensions, power):
    left, singular = randomized_svd(matrix, dimensions)
    vectors = left * (singular**power)
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    np.divide(vectors, norms, out=vectors, where=norms > 0)
    return vectors

def cosine_neighbors(matrix, limit, block=512):
    size = matrix.shape[0]
    found = {}
    for start in range(0, size, block):
        stop = min(start + block, size)
        scores = matrix[start:stop] @ matrix.T
        for offset in range(stop - start):
            target = start + offset
            row = scores[offset]
            row[target] = -1.0
            top = np.argpartition(-row, limit)[:limit]
            top = top[np.argsort(-row[top])]
            found[target] = [
                (int(other), float(row[other])) for other in top if row[other] > 0
            ]
    return found

COARSE = {
    "N": "N",
    "V": "V",
    "Vi": "V",
    "Vp": "V",
    "Vpt": "V",
    "Vst": "V",
    "Vaux": "V",
    "V-sep": "V",
    "Vp-sep": "V",
    "Vs": "A",
    "Vs-attr": "A",
    "Vs-pred": "A",
    "Vs-sep": "A",
    "Adv": "D",
    "M": "M",
}

def parts_of_speech_table():
    path = os.path.join(DATA, "huayu", "parts_of_speech.json")
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as handle:
        return json.load(handle)

def coarse_classes(vocabulary, table):
    classes = {}
    for number, word in enumerate(vocabulary):
        tags = re.split(r"[/／]", (table.get(word) or {}).get("tocfl") or "")
        found = {COARSE[tag.strip()] for tag in tags if tag.strip() in COARSE}
        if found:
            classes[number] = found
    return classes

def function_words(vocabulary, table):
    closed = set()
    for number, word in enumerate(vocabulary):
        if word in FUNCTION_WORDS:
            closed.add(number)
            continue
        tags = re.split(r"[/／]", (table.get(word) or {}).get("tocfl") or "")
        tags = [tag.strip() for tag in tags if tag.strip()]
        if tags and set(tags) <= FUNCTION_POS:
            closed.add(number)
    return closed

def log_dice(observed, left, right):
    return 14 + np.log2(2 * observed / (left + right))

def log_likelihood(keys, counts, size, min_count):
    marginals = np.bincount(
        keys // size, weights=counts.astype(np.float64), minlength=size
    )
    total_pairs = float(counts.sum())

    keep = counts >= min_count
    keys = keys[keep]
    observed = counts[keep].astype(np.float64)
    first = keys // size
    second = keys % size

    left = marginals[first]
    right = marginals[second]
    expected = left * right / total_pairs

    a = observed
    b = left - a
    c = right - a
    d = total_pairs - a - b - c

    def term(observed_cell, expected_cell):
        with np.errstate(divide="ignore", invalid="ignore"):
            value = observed_cell * np.log(observed_cell / expected_cell)
        return np.where(observed_cell > 0, value, 0.0)

    e_a = expected
    e_b = left * (total_pairs - right) / total_pairs
    e_c = (total_pairs - left) * right / total_pairs
    e_d = (total_pairs - left) * (total_pairs - right) / total_pairs

    g2 = 2 * (term(a, e_a) + term(b, e_b) + term(c, e_c) + term(d, e_d))
    dice = log_dice(observed, left, right)
    attracted = (observed > expected) & (g2 >= COLLOCATE_MIN_G2)
    return first[attracted], second[attracted], dice[attracted]

def evaluate(found, index, synonyms, at):
    hits = 0
    total = 0
    reciprocal = 0.0
    covered = 0
    for word, partners in synonyms.items():
        target = index.get(word)
        if target is None or target not in found:
            continue
        gold = {index[partner] for partner in partners if partner in index}
        if not gold:
            continue
        covered += 1
        ranked = [other for other, _ in found[target][:at]]
        total += len(gold)
        hits += len(gold & set(ranked))
        for position, other in enumerate(ranked, start=1):
            if other in gold:
                reciprocal += 1 / position
                break
    return {
        "words with gold standard": covered,
        "recall": round(hits / total, 4) if total else 0.0,
        "MRR": round(reciprocal / covered, 4) if covered else 0.0,
    }

def main():
    synonyms, antonyms = dictionary_relations()
    print(
        f"dictionary synonyms : {len(synonyms)} words, {sum(len(v) for v in synonyms.values())} pairs"
    )
    print(
        f"dictionary antonyms : {len(antonyms)} words, {sum(len(v) for v in antonyms.values())} pairs"
    )

    words = dictionary_words()
    raw_stream, alphabet = cached_stream(words)
    frequency = np.bincount(raw_stream[raw_stream >= 0], minlength=len(alphabet))
    print(f"corpus tokens      : {int(frequency.sum())}, distinct {len(alphabet)}")

    vocabulary = [
        alphabet[number] for number in np.flatnonzero(frequency >= MIN_TARGET_FREQ)
    ]
    frequency_of = frequency[frequency >= MIN_TARGET_FREQ]
    index = {word: number for number, word in enumerate(vocabulary)}
    size = len(vocabulary)
    print(f"working vocabulary : {size} words (at least {MIN_TARGET_FREQ})")

    stream = restrict(raw_stream, alphabet, index)

    keys, pair_totals = pair_counts(stream, size, WINDOW)
    print(f"pairs in window +-{WINDOW} : {len(keys)}")

    if os.environ.get("SWEEP"):
        for window in (2, 3):
            window_keys, window_totals = pair_counts(stream, size, window)
            targets, contexts, values = sppmi_rows(
                window_keys, window_totals, size, ALPHA, SHIFT, MIN_PAIR_COUNT
            )
            matrix = build_space(targets, contexts, values, size, TOP_CONTEXTS)
            score = evaluate(
                cosine_neighbors(matrix, NEIGHBORS), index, synonyms, NEIGHBORS
            )
            print(f"  window +-{window} no SVD: {score}", flush=True)
            for dimensions in (200, 300, 500):
                for power in (0.0, 0.5, 1.0):
                    vectors = embed(matrix, dimensions, power)
                    score = evaluate(
                        cosine_neighbors(vectors, NEIGHBORS),
                        index,
                        synonyms,
                        NEIGHBORS,
                    )
                    print(
                        f"  window +-{window} SVD d={dimensions} p={power}: {score}",
                        flush=True,
                    )
                    del vectors
            del matrix, window_keys, window_totals
        return

    targets, contexts, values = sppmi_rows(
        keys, pair_totals, size, ALPHA, SHIFT, MIN_PAIR_COUNT
    )
    matrix = build_space(targets, contexts, values, size, TOP_CONTEXTS)
    found = cosine_neighbors(matrix, MUTUAL_RANK)
    del matrix
    print(
        f"quality (a={ALPHA}, k={SHIFT}): {evaluate(found, index, synonyms, NEIGHBORS)}"
    )

    rank_of = {
        target: {other: position for position, (other, _) in enumerate(rows)}
        for target, rows in found.items()
    }

    collocate_keys, collocate_counts = pair_counts(stream, size, COLLOCATE_WINDOW)
    left, right, dice = log_likelihood(
        collocate_keys, collocate_counts, size, COLLOCATE_MIN_COUNT
    )
    table = parts_of_speech_table()
    closed = function_words(vocabulary, table)
    classes = coarse_classes(vocabulary, table)
    rare = {
        number
        for number in range(size)
        if frequency_of[number] < COLLOCATE_MIN_PARTNER_FREQ
    }
    strong = collections.defaultdict(list)
    for first, second, value in zip(left.tolist(), right.tolist(), dice.tolist()):
        if second not in closed and second not in rare:
            strong[first].append((second, value))
    for rows in strong.values():
        rows.sort(key=lambda item: -item[1])
    print(
        f"collocations G2>={COLLOCATE_MIN_G2}, function words excluded: {sum(len(rows) for rows in strong.values())}"
    )

    payload = {}
    for word in set(list(synonyms) + list(antonyms) + vocabulary):
        entry = {}
        known_synonyms = synonyms.get(word, [])
        known_antonyms = antonyms.get(word, [])
        if known_synonyms:
            entry["synonyms"] = known_synonyms
        if known_antonyms:
            entry["antonyms"] = known_antonyms

        target = index.get(word)
        if target is not None:
            taken = set(known_synonyms) | set(known_antonyms)
            related = []
            for other, score in found.get(target, []):
                partner = vocabulary[other]
                if score < MIN_COSINE or len(related) >= NEIGHBORS:
                    break
                if rank_of.get(other, {}).get(target, MUTUAL_RANK) >= MUTUAL_RANK:
                    continue
                if partner in taken or partner in word or word in partner:
                    continue
                if other in closed:
                    continue
                mine, theirs = classes.get(target), classes.get(other)
                if mine and theirs and not (mine & theirs):
                    continue
                related.append({"word": partner, "score": round(score, 3)})
            if related:
                entry["related"] = related

            collocates = []
            for other, value in strong.get(target, []):
                partner = vocabulary[other]
                if len(collocates) >= COLLOCATES:
                    break
                if partner in word or word in partner:
                    continue
                collocates.append({"word": partner, "dice": round(value, 2)})
            if collocates:
                entry["collocates"] = collocates

        if entry:
            payload[word] = entry

    with open(OUT, "w", encoding="utf-8") as handle:
        json.dump(dict(sorted(payload.items())), handle, ensure_ascii=False)

    print(f"words in thesaurus : {len(payload)}")
    print(
        f"  with related     : {sum(1 for e in payload.values() if e.get('related'))}"
    )
    print(
        f"  with collocates  : {sum(1 for e in payload.values() if e.get('collocates'))}"
    )
    print(f"-> {OUT}")

if __name__ == "__main__":
    main()
